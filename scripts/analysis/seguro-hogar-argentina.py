"""Cálculos propios del caso seguro-hogar-argentina sobre fuentes SSN.

Reproduce cada número con estatus "cálculo propio sobre balances/anexo SSN":
prima emitida del subramo Combinado Familiar (el seguro de hogar), su
participación de mercado, sus anulaciones, el top de aseguradoras y el rank de
La Caja, más la variación interanual de pólizas vigentes del ramo contra el
total del mercado (el dato que decide el gancho del caso). El auditor del nodo 5
corre este script y compara la salida contra la hoja de hechos (regla 2.6).

Aislar el hogar del resto del ramo importa: la SSN agrupa "Combinado Familiar e
Integral", pero el subramo 1.020.02 es Integral de Comercio (comercios, no
hogar). Trabajar el ramo entero mezclaría hogar con comercio (la clase de doble
conteo de la falla #7 del registro). Por eso los balances se leen a nivel
subramo 1.020.01. El anexo de pólizas, en cambio, solo publica el ramo agregado
"Combinado Familiar e Integ.", y así se cita.

Datos de entrada (no viajan en el repo, ver .gitignore):
  data/seguro-hogar-argentina/balances.csv
    Balances trimestrales de aseguradoras, SSN datos abiertos:
    https://datosabiertos.ssn.gob.ar/dataset/balances
    Corte usado: marzo 2026 (importes acumulados del ejercicio jul-2025 a mar-2026).
  data/seguro-hogar-argentina/ssn_polizas_1t2026.xlsx
    Anexo estadístico de pólizas del 1º trimestre 2026:
    https://www.argentina.gob.ar/sites/default/files/ssn_202603_polizas_siniestros.xlsx
  data/seguro-hogar-argentina/ssn_polizas_1t2025.xlsx
    Anexo estadístico de pólizas del 1º trimestre 2025:
    https://www.argentina.gob.ar/sites/default/files/ssn_202501_polizas_siniestros.xlsx

Correr desde la raíz del repo:  python3 scripts/analysis/seguro-hogar-argentina.py
"""

import csv
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
DATA = ROOT / "data" / "seguro-hogar-argentina"

# Cuentas a nivel subramo del plan contable SSN. El detalle por subramo evita el
# doble conteo con las cuentas agregadas (la falla #7 del registro del flujo).
PRIMAS = "5.01.01.01.01.01.01.00"  # Primas Seguros Directos
ANUL = "4.01.04.04.04.01.01.00"  # Anulaciones de Primas - Seguros Directos
SUB_HOGAR = "1.020.01"  # Comb. Fam. e Int. - Comb. Fam. (el seguro de hogar)
SUB_INTCOM = "1.020.02"  # Comb. Fam. e Int. - Int. de Com. (integral de comercio)


def num(s: str) -> float:
    s = (s or "").strip()
    if s in ("", "-"):
        return 0.0
    neg = s.startswith("(") or s.startswith("-")
    s = s.strip("()-").replace(".", "").replace(",", ".")
    try:
        v = float(s)
    except ValueError:
        return 0.0
    return -v if neg else v


def balances() -> None:
    tot = hogar = hogar_anul = intcom = 0.0
    cias: dict[str, float] = {}
    with open(DATA / "balances.csv", encoding="utf-8", errors="replace") as f:
        for row in csv.DictReader(f):
            if not row["subramo_id"]:
                continue
            v = num(row["importe"])
            c = row["cuenta_id"]
            if c == PRIMAS:
                tot += v
                if row["subramo_id"] == SUB_HOGAR:
                    hogar += v
                    cias[row["cia_denominacion"]] = cias.get(row["cia_denominacion"], 0) + v
                elif row["subramo_id"] == SUB_INTCOM:
                    intcom += v
            elif c == ANUL and row["subramo_id"] == SUB_HOGAR:
                hogar_anul += v

    print("== Balances SSN (jul-2025 a mar-2026, seguros directos por subramo) ==")
    print(f"Prima emitida total del mercado:  ${tot / 1e12:,.2f} billones")
    print(f"Seguro de hogar (subramo 1.020.01): ${hogar / 1e9:,.1f} mil M  ({hogar / tot * 100:.2f}% del mercado)")
    print(f"  (referencia: Integral de Comercio 1.020.02: ${intcom / 1e9:,.1f} mil M, {intcom / tot * 100:.2f}%)")
    print(f"Anulaciones del hogar:            ${abs(hogar_anul) / 1e9:,.1f} mil M  ({abs(hogar_anul) / hogar * 100:.1f}% de su prima)")
    top = sorted(cias.items(), key=lambda x: -x[1])[:10]
    top3 = sum(v for _, v in top[:3])
    print(f"Top 3 del hogar:                  {top3 / hogar * 100:.1f}% de la prima")
    print("Top 10 del hogar:")
    for i, (cia, v) in enumerate(top, 1):
        print(f"  {i:2} {cia[:44]:46} ${v / 1e9:7,.1f} mil M  {v / hogar * 100:5.1f}%")
    caja = next(((i, c, v) for i, (c, v) in enumerate(sorted(cias.items(), key=lambda x: -x[1]), 1) if c == "CAJA DE SEGUROS S.A."), None)
    if caja:
        print(f"La Caja (Caja de Seguros S.A.): rank #{caja[0]}, ${caja[2] / 1e9:,.1f} mil M ({caja[2] / hogar * 100:.1f}%)")


def _annex_rows(path: Path) -> dict:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True)
    # el nombre de la hoja cambió entre entregas ("Cuadro 1 " vs "Cuadros 1 ")
    sheet = next(s for s in wb.sheetnames if s.strip().startswith("Cuadro"))
    ws = wb[sheet]
    out = {}
    for row in ws.iter_rows(values_only=True):
        name = str(row[0] or "").strip()
        if name:
            out[name] = row
    return out


def _get(rows: dict, key: str):
    for name, row in rows.items():
        if key.lower() in name.lower():
            return name, row
    return None, None


def polizas() -> None:
    # Columnas del Cuadro 1: [1] emitidas netas, [5] vigentes al cierre,
    # [6] part. vigentes, [7] var. interanual de vigentes (publicada por la SSN).
    r25 = _annex_rows(DATA / "ssn_polizas_1t2025.xlsx")
    r26 = _annex_rows(DATA / "ssn_polizas_1t2026.xlsx")
    print("\n== Anexo estadístico SSN, Cuadro 1: pólizas vigentes interanual ==")
    print(f"{'Ramo':30} {'1T2025':>12} {'1T2026':>12} {'calc':>8} {'SSN pub':>8}")
    for key in ("Combinado Familiar", "Incendio", "Robo", "T O T A L"):
        n25, a = _get(r25, key)
        n26, b = _get(r26, key)
        if not a or not b:
            continue
        v25, v26 = a[5], b[5]
        calc = (v26 - v25) / v25 * 100
        pub = b[7]
        print(f"{n26[:30]:30} {v25:>12,} {v26:>12,} {calc:>7.2f}% {pub:>7.2f}%")
    _, tot26 = _get(r26, "T O T A L")
    _, c25 = _get(r25, "Combinado Familiar")
    _, c26 = _get(r26, "Combinado Familiar")
    print(f"\nParticipación del hogar en pólizas vigentes: {c25[6]:.2f}% (1T2025) -> {c26[6]:.2f}% (1T2026)")
    print("Lectura: el hogar queda plano mientras el mercado crece, así que pierde participación.")


# INDEC, Censo Nacional de Población, Hogares y Viviendas 2022, resultados
# definitivos: viviendas particulares ocupadas. Denominador correcto para
# penetración (una póliza de hogar asegura una vivienda).
# https://censo.gob.ar/index.php/datos_definitivos_total_pais/
VIVIENDAS_OCUPADAS = 15_699_016


def penetracion() -> None:
    # Vigentes del ramo agregado (incluye Integral de Comercio): techo, porque
    # el numerador tiene comercios que no son hogares.
    r26 = _annex_rows(DATA / "ssn_polizas_1t2026.xlsx")
    _, c26 = _get(r26, "Combinado Familiar")
    vig_ramo = c26[5]

    # Share del hogar dentro del ramo, medido por prima en los balances (subramo).
    tot_ramo_prima = hogar_prima = 0.0
    with open(DATA / "balances.csv", encoding="utf-8", errors="replace") as f:
        for row in csv.DictReader(f):
            if row["cuenta_id"] == PRIMAS and row["subramo_id"] in (SUB_HOGAR, SUB_INTCOM, "1.020.99"):
                v = num(row["importe"])
                tot_ramo_prima += v
                if row["subramo_id"] == SUB_HOGAR:
                    hogar_prima += v
    share = hogar_prima / tot_ramo_prima

    print("\n== Penetración del seguro de hogar (cálculo propio SSN x INDEC) ==")
    print(f"Denominador: {VIVIENDAS_OCUPADAS:,} viviendas particulares ocupadas (INDEC 2022).")
    print(f"Techo (ramo entero, incluye comercio): {vig_ramo / VIVIENDAS_OCUPADAS * 100:.1f}%")
    print(f"Share del hogar en el ramo por prima: {share * 100:.1f}% (hipótesis: pólizas ~ prima)")
    for s, lbl in [(share, "pólizas ~ prima"), (0.60, "hogar 60% de pólizas"), (0.70, "hogar 70%")]:
        vh = vig_ramo * s
        print(f"  hogar ~{vh / 1e6:.2f}M pólizas ({lbl}): {vh / VIVIENDAS_OCUPADAS * 100:.1f}% de las viviendas")
    print("Contraste: Chubb (encuesta autorreportada) dice 45%; imposible contra el techo de 25%.")


if __name__ == "__main__":
    balances()
    polizas()
    penetracion()
